"""
search_glossary.py - Search business glossary Excel files for terms and definitions.

Reads .xlsx/.xls business glossary files (e.g. from documents/ or a dedicated folder),
searches for keywords across sheet names, column headers, and cell values.
Caches parsed data as JSON for fast repeat searches.

Business glossary columns (Vietnamese):
  STT; BG_CODE; Tên chỉ tiêu/thuật ngữ; Tên chỉ tiêu/thuật ngữ (đề xuất);
  Chiều filter; Domain; Phân loại; Đơn vị sở hữu; Định Nghĩa; Cách tính;
  Mục đích sử dụng; Thứ tự ưu tiên; Đơn vị đưa ra định nghĩa; Đơn vị thực hiện tính toán;
  Ghi chú; PIC_DA; Bảng dữ liệu DWH; Trường dữ liệu DWH; SQL tính toán;
  Nguồn dữ liệu gốc; Bảng dữ liệu gốc; Trường dữ liệu gốc

Usage:
    python search_glossary.py --keyword "doanh thu"
    python search_glossary.py --keyword "DWH" --folder documents/
    python search_glossary.py --keyword "định nghĩa" --no-cache
    python search_glossary.py --keyword "REVENUE" --format json
"""

import argparse
import hashlib
import json
import os
import sys
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None


# ============================================================================
# Cache
# ============================================================================

CACHE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".doc_cache")


def _file_hash(filepath: str) -> str:
    """Return MD5 of file content for cache invalidation."""
    h = hashlib.md5()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def _cache_path(filepath: str) -> str:
    name = Path(filepath).stem
    return os.path.join(CACHE_DIR, f"{name}.json")


def _load_cache(filepath: str) -> list[dict] | None:
    """Load cached parsed data if still valid."""
    cp = _cache_path(filepath)
    if not os.path.exists(cp):
        return None
    try:
        with open(cp, "r", encoding="utf-8") as f:
            cached = json.load(f)
        if cached.get("file_hash") == _file_hash(filepath):
            return cached["rows"]
    except (json.JSONDecodeError, KeyError):
        pass
    return None


def _save_cache(filepath: str, rows: list[dict]):
    os.makedirs(CACHE_DIR, exist_ok=True)
    cp = _cache_path(filepath)
    with open(cp, "w", encoding="utf-8") as f:
        json.dump({"file_hash": _file_hash(filepath), "rows": rows}, f, ensure_ascii=False)


# ============================================================================
# Business glossary: doc type and column mappings
# ============================================================================

DOC_TYPE_GLOSSARY = "glossary"


def _is_glossary_file(filename: str) -> bool:
    """Recognize business glossary files by name pattern."""
    base = Path(filename).stem.lower()
    return (
        "glossary" in base
        or "business-glossary" in base
        or "bg-" in base
        or base == "business_glossary"
    )


def _normalize_header(h: str) -> str:
    return (h or "").strip().lower()


# Map known Vietnamese headers (normalized) to canonical key for glossary
# Columns: STT; BG_CODE; Tên chỉ tiêu/thuật ngữ; Tên chỉ tiêu/thuật ngữ (đề xuất);
# Chiều filter; Domain; Phân loại; Đơn vị sở hữu; Định Nghĩa; Cách tính;
# Mục đích sử dụng; Thứ tự ưu tiên; Đơn vị đưa ra định nghĩa; Đơn vị thực hiện tính toán;
# Ghi chú; PIC_DA; Bảng dữ liệu DWH; Trường dữ liệu DWH; SQL tính toán;
# Nguồn dữ liệu gốc; Bảng dữ liệu gốc; Trường dữ liệu gốc
GLOSSARY_HEADER_MAP = {
    "stt": "stt",
    "bg_code": "bg_code",
    "tên chỉ tiêu/thuật ngữ": "term_name",
    "tên chỉ tiêu/thuật ngữ (đề xuất)": "term_name_proposed",
    "chiều filter": "filter_dimension",
    "domain": "domain",
    "phân loại": "category",
    "đơn vị sở hữu": "owner_unit",
    "định nghĩa": "definition",
    "cách tính": "calculation",
    "mục đích sử dụng": "usage_purpose",
    "thứ tự ưu tiên (5 is highest)": "priority",
    "thứ tự ưu tiên": "priority",
    "đơn vị đưa ra định nghĩa": "definition_unit",
    "đơn vị thực hiện tính toán": "calculation_unit",
    "ghi chú": "note",
    "pic_da": "pic_da",
    "bảng dữ liệu dwh": "dwh_table",
    "trường dữ liệu dwh": "dwh_column",
    "sql tính toán": "sql_calculation",
    "nguồn dữ liệu gốc": "source_system",
    "bảng dữ liệu gốc": "source_table",
    "trường dữ liệu gốc": "source_column",
}


def _header_to_canonical(header: str) -> str | None:
    n = _normalize_header(header)
    if not n:
        return None
    return GLOSSARY_HEADER_MAP.get(n)


def _build_header_index(headers: list[str]) -> dict[str, int]:
    """Map canonical key -> column index (first match)."""
    idx = {}
    for i, h in enumerate(headers):
        c = _header_to_canonical(h)
        if c and c not in idx:
            idx[c] = i
    return idx


def _row_to_normalized_context(row: dict) -> str:
    """
    Build a readable context string for a glossary row.
    Prioritizes: term name, definition, calculation, DWH table/column, SQL.
    """
    headers = row.get("headers") or []
    values = row.get("values") or []
    idx = _build_header_index(headers)
    parts = []

    def _v(key: str, max_len: int = 80) -> str:
        if key not in idx or idx[key] >= len(values):
            return ""
        v = (values[idx[key]] or "").strip()
        if not v:
            return ""
        return (v[:max_len] + "…") if len(v) > max_len else v

    if _v("term_name"):
        parts.append(f"Term={_v('term_name', 50)}")
    if _v("definition"):
        parts.append(f"Definition={_v('definition')}")
    if _v("calculation"):
        parts.append(f"Calculation={_v('calculation')}")
    if _v("dwh_table"):
        parts.append(f"DWH_Table={_v('dwh_table', 40)}")
    if _v("dwh_column"):
        parts.append(f"DWH_Column={_v('dwh_column', 40)}")
    if _v("sql_calculation"):
        parts.append(f"SQL={_v('sql_calculation', 60)}")
    if _v("domain"):
        parts.append(f"Domain={_v('domain', 30)}")
    if _v("source_table"):
        parts.append(f"Source_Table={_v('source_table', 40)}")

    if parts:
        return " | ".join(parts)

    # Fallback: first 6 non-empty key=value
    context_parts = []
    for i, val in enumerate(values[:14]):
        if val and str(val).strip():
            hdr = headers[i] if i < len(headers) else f"col_{i}"
            context_parts.append(f"{hdr}={str(val).strip()[:50]}")
    return " | ".join(context_parts[:6])


def _row_glossary_extra(row: dict) -> dict:
    """Extract glossary-specific fields for result entry (table_name, column_name, term_name, etc.)."""
    headers = row.get("headers") or []
    values = row.get("values") or []
    idx = _build_header_index(headers)
    out = {}

    def _v(key: str):
        if key not in idx or idx[key] >= len(values):
            return None
        v = (values[idx[key]] or "").strip()
        return v or None

    if _v("term_name"):
        out["term_name"] = _v("term_name")
    if _v("bg_code"):
        out["bg_code"] = _v("bg_code")
    if _v("dwh_table"):
        out["table_name"] = _v("dwh_table")
    if _v("dwh_column"):
        out["column_name"] = _v("dwh_column")
    if _v("definition"):
        out["definition"] = _v("definition")
    if _v("sql_calculation"):
        out["sql_calculation"] = _v("sql_calculation")
    if _v("domain"):
        out["domain"] = _v("domain")
    return out


# ============================================================================
# Parsers
# ============================================================================

def _parse_xlsx(filepath: str) -> list[dict]:
    """Parse .xlsx file into list of {file, sheet, row_num, headers, values}."""
    if openpyxl is None:
        print("openpyxl is required for .xlsx files. Install: pip install openpyxl", file=sys.stderr)
        return []

    rows = []
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            str_values = [str(v) if v is not None else "" for v in row]
            if row_idx == 1:
                headers = str_values
                continue
            rows.append({
                "file": os.path.basename(filepath),
                "sheet": sheet_name,
                "row_num": row_idx,
                "headers": headers,
                "values": str_values,
            })
    wb.close()
    return rows


def _parse_xls(filepath: str) -> list[dict]:
    """Parse .xls file into list of dicts."""
    if xlrd is None:
        print("xlrd is required for .xls files. Install: pip install xlrd", file=sys.stderr)
        return []

    rows = []
    wb = xlrd.open_workbook(filepath)
    for sheet in wb.sheets():
        headers = []
        for row_idx in range(sheet.nrows):
            str_values = [str(sheet.cell_value(row_idx, c)) for c in range(sheet.ncols)]
            if row_idx == 0:
                headers = str_values
                continue
            rows.append({
                "file": os.path.basename(filepath),
                "sheet": sheet.name,
                "row_num": row_idx + 1,
                "headers": headers,
                "values": str_values,
            })
    return rows


def parse_file(filepath: str, use_cache: bool = True) -> list[dict]:
    """Parse an Excel file (with optional caching)."""
    if use_cache:
        cached = _load_cache(filepath)
        if cached is not None:
            return cached

    ext = Path(filepath).suffix.lower()
    if ext == ".xlsx":
        rows = _parse_xlsx(filepath)
    elif ext == ".xls":
        rows = _parse_xls(filepath)
    else:
        print(f"Unsupported file format: {ext}", file=sys.stderr)
        return []

    if use_cache:
        _save_cache(filepath, rows)

    return rows


# ============================================================================
# Search
# ============================================================================

def search_glossary(keyword: str, folder: str = "documents/",
                   use_cache: bool = True, use_regex: bool = False,
                   limit: int = 200, glossary_only: bool = True) -> list[dict]:
    """
    Search business glossary (and optionally all) Excel files in folder for keyword.

    If glossary_only=True, only files matching glossary naming are searched.
    Returns list of matches with context: file, sheet, row_num, matched_field, context,
    plus term_name, table_name, column_name when available.
    """
    skill_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    abs_folder = os.path.join(skill_root, folder) if not os.path.isabs(folder) else folder

    if not os.path.isdir(abs_folder):
        print(f"Thư mục không tồn tại: {abs_folder}", file=sys.stderr)
        return []

    excel_files = []
    for f in os.listdir(abs_folder):
        if f.lower().endswith((".xlsx", ".xls")) and not f.startswith("~$"):
            if not glossary_only or _is_glossary_file(f):
                excel_files.append(os.path.join(abs_folder, f))

    if not excel_files:
        msg = "glossary" if glossary_only else "Excel"
        print(f"Không tìm thấy file {msg} trong: {abs_folder}", file=sys.stderr)
        return []

    results = []
    for fp in sorted(excel_files):
        rows = parse_file(fp, use_cache=use_cache)

        seen_sheets = set()
        for row in rows:
            sheet = row["sheet"]
            if sheet not in seen_sheets and _match(keyword, sheet, use_regex):
                seen_sheets.add(sheet)
                results.append({
                    "file": row["file"], "sheet": sheet,
                    "row_num": None, "matched_field": "sheet_name",
                    "context": f"Sheet name: {sheet}",
                    "doc_type": DOC_TYPE_GLOSSARY,
                })

        seen_headers = set()
        for row in rows:
            sheet = row["sheet"]
            for hi, h in enumerate(row["headers"]):
                key = (sheet, hi)
                if key not in seen_headers and _match(keyword, h, use_regex):
                    seen_headers.add(key)
                    results.append({
                        "file": row["file"], "sheet": sheet,
                        "row_num": 1, "matched_field": "header",
                        "context": f"Header [{hi}]: {h}",
                        "doc_type": DOC_TYPE_GLOSSARY,
                    })

        for row in rows:
            for vi, v in enumerate(row["values"]):
                if _match(keyword, v, use_regex):
                    header = row["headers"][vi] if vi < len(row["headers"]) else f"col_{vi}"
                    ctx = _row_to_normalized_context(row)
                    extra = _row_glossary_extra(row)

                    entry = {
                        "file": row["file"], "sheet": row["sheet"],
                        "row_num": row["row_num"],
                        "matched_field": f"cell:{header}",
                        "context": ctx,
                        "doc_type": DOC_TYPE_GLOSSARY,
                    }
                    entry.update(extra)
                    results.append(entry)

                if len(results) >= limit:
                    break
            if len(results) >= limit:
                break
        if len(results) >= limit:
            break

    return results[:limit]


def _match(pattern: str, text: str, use_regex: bool) -> bool:
    if not text:
        return False
    if use_regex:
        return bool(re.search(pattern, text, re.IGNORECASE))
    return pattern.lower() in text.lower()


# ============================================================================
# Formatters
# ============================================================================

def format_text(results: list[dict]) -> str:
    if not results:
        return "Không tìm thấy kết quả nào."

    lines = [f"Tìm thấy {len(results)} kết quả (Business Glossary):\n"]
    lines.append(f"{'FILE':<28} {'SHEET':<18} {'ROW':<5} {'MATCH':<14} {'CONTEXT'}")
    lines.append("-" * 130)
    for r in results:
        lines.append(
            f"{r['file']:<28} {r['sheet']:<18} {str(r['row_num'] or ''):<5} "
            f"{r['matched_field']:<14} {(r['context'] or '')[:85]}"
        )
    return "\n".join(lines)


def format_json(results: list[dict]) -> str:
    return json.dumps(results, indent=2, ensure_ascii=False)


def format_markdown(results: list[dict]) -> str:
    if not results:
        return "Không tìm thấy kết quả nào."
    lines = [f"Tìm thấy **{len(results)}** kết quả (Business Glossary):\n"]
    lines.append("| File | Sheet | Row | Match | Context |")
    lines.append("|------|-------|-----|-------|---------|")
    for r in results:
        ctx = (r["context"] or "")[:80].replace("|", "\\|")
        lines.append(
            f"| {r['file']} | {r['sheet']} | {r['row_num'] or ''} "
            f"| {r['matched_field']} | {ctx} |"
        )
    return "\n".join(lines)


# ============================================================================
# CLI
# ============================================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Search business glossary Excel files for terms and definitions."
    )
    parser.add_argument("--keyword", "-k", required=True, help="Search keyword or regex")
    parser.add_argument("--folder", default="documents/", help="Folder containing glossary Excel files")
    parser.add_argument("--no-cache", action="store_true", help="Disable caching")
    parser.add_argument("--regex", action="store_true", help="Use regex matching")
    parser.add_argument("--limit", type=int, default=200, help="Max results (default: 200)")
    parser.add_argument("--all-excel", action="store_true",
                        help="Search all Excel files in folder, not only glossary-named files")
    parser.add_argument("--format", choices=["text", "json", "markdown"], default="text")
    args = parser.parse_args()

    try:
        results = search_glossary(
            keyword=args.keyword, folder=args.folder,
            use_cache=not args.no_cache, use_regex=args.regex, limit=args.limit,
            glossary_only=not args.all_excel,
        )
        if args.format == "json":
            print(format_json(results))
        elif args.format == "markdown":
            print(format_markdown(results))
        else:
            print(format_text(results))
    except Exception as e:
        print(f"Lỗi: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)
