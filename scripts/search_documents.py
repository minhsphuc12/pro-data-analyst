"""
search_documents.py - Search Excel documentation files for table/column info.

Reads .xlsx/.xls files from the documents/ folder, searches for keywords across
sheet names, column headers, and cell values.  Caches parsed data as JSON for
fast repeat searches.

Supports two documentation types:
- DWH: dwh-meta-tables.xlsx, dwh-meta-columns.xlsx (Vietnamese headers)
- Source systems: [source]-meta-tables.xlsx, [source]-meta-columns.xlsx (English headers)

Usage:
    python search_documents.py --keyword "khach hang"
    python search_documents.py --keyword "doanh thu" --folder documents/
    python search_documents.py --keyword "REVENUE" --no-cache
    python search_documents.py --keyword "CONTRACT" --format json
"""

import argparse
import hashlib
import json
import os
import sys
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None


# ============================================================================
# Cache
# ============================================================================

CACHE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".doc_cache")


def _file_hash(filepath: str) -> str:
    """Return MD5 of file content for cache invalidation."""
    h = hashlib.md5()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def _cache_path(filepath: str) -> str:
    name = Path(filepath).stem
    return os.path.join(CACHE_DIR, f"{name}.json")


def _load_cache(filepath: str) -> list[dict] | None:
    """Load cached parsed data if still valid."""
    cp = _cache_path(filepath)
    if not os.path.exists(cp):
        return None
    try:
        with open(cp, "r", encoding="utf-8") as f:
            cached = json.load(f)
        if cached.get("file_hash") == _file_hash(filepath):
            return cached["rows"]
    except (json.JSONDecodeError, KeyError):
        pass
    return None


def _save_cache(filepath: str, rows: list[dict]):
    os.makedirs(CACHE_DIR, exist_ok=True)
    cp = _cache_path(filepath)
    with open(cp, "w", encoding="utf-8") as f:
        json.dump({"file_hash": _file_hash(filepath), "rows": rows}, f, ensure_ascii=False)


# ============================================================================
# Document type detection and column mappings
# ============================================================================

# Doc type: "dwh_tables" | "dwh_columns" | "source_tables" | "source_columns" | None (generic)
def _doc_type_from_filename(filename: str) -> tuple[str | None, str | None]:
    """
    Return (doc_type, source_name). source_name is set for source_* types (e.g. "cif", "appraisal").
    """
    base = Path(filename).stem.lower()
    if base == "dwh-meta-tables":
        return "dwh_tables", None
    if base == "dwh-meta-columns":
        return "dwh_columns", None
    if base.endswith("-meta-tables"):
        return "source_tables", base.replace("-meta-tables", "")
    if base.endswith("-meta-columns"):
        return "source_columns", base.replace("-meta-columns", "")
    return None, None


# Header name (normalized: strip, lower) -> canonical key for context
# DWH tables: Tên Bảng- dwh, Mô tả bảng, Schema, Source, Domain, ...
# DWH columns: TÊN BẢNG, TÊN TRƯỜNG, MÔ TẢ TRƯỜNG, KIỂU DỮ LIỆU, ...
# Source tables: Table Name, Description, Care, Type
# Source columns: Column Name, Data Type, Comment, Sample Data, Table Name

def _normalize_header(h: str) -> str:
    return (h or "").strip().lower()


# Map known headers (normalized) to canonical key
def _header_to_canonical(header: str, doc_type: str | None) -> str | None:
    n = _normalize_header(header)
    if not n:
        return None
    # DWH: Vietnamese headers
    if doc_type in ("dwh_tables", "dwh_columns", None):
        if n in ("tên bảng- dwh", "tên bảng", "ten bang", "table name"):
            return "table_name"
        if n in ("tên trường", "tên trưòng", "ten truong", "column name"):
            return "column_name"
        if n in ("mô tả bảng", "mô tả trường", "mo ta bang", "mo ta truong", "description", "comment"):
            return "description"
        if n in ("kiểu dữ liệu", "kieu du lieu", "data type"):
            return "data_type"
        if n in ("schema", "source", "domain", "hệ thống source", "he thong source"):
            return "source" if "source" in n or n == "source" else n
    # Source docs: English headers (Table Name, Description, Column Name, Data Type, Comment, Sample Data)
    if doc_type in ("source_tables", "source_columns", None):
        if n == "table name":
            return "table_name"
        if n == "column name":
            return "column_name"
        if n in ("description", "comment"):
            return "description"
        if n == "data type":
            return "data_type"
        if n in ("care", "type"):
            return n
        if n == "sample data":
            return "sample_data"
    return None


def _build_header_index(headers: list[str], doc_type: str | None) -> dict[str, int]:
    """Map canonical key -> column index (first match)."""
    idx = {}
    for i, h in enumerate(headers):
        c = _header_to_canonical(h, doc_type)
        if c and c not in idx:
            idx[c] = i
    return idx


def _row_to_normalized_context(row: dict, doc_type: str | None, source_name: str | None) -> str:
    """
    Build a readable context string for a row using known column semantics.
    Falls back to key=value for first 8 fields if not recognized.
    """
    headers = row.get("headers") or []
    values = row.get("values") or []
    if doc_type and headers:
        idx = _build_header_index(headers, doc_type)
        parts = []
        if "table_name" in idx:
            v = values[idx["table_name"]] if idx["table_name"] < len(values) else ""
            if v:
                parts.append(f"Table={v}")
        if "column_name" in idx and doc_type in ("dwh_columns", "source_columns"):
            v = values[idx["column_name"]] if idx["column_name"] < len(values) else ""
            if v:
                parts.append(f"Column={v}")
        if "description" in idx:
            v = values[idx["description"]] if idx["description"] < len(values) else ""
            if v:
                parts.append(f"Description={v[:60]}{'…' if len(v) > 60 else ''}")
        if "data_type" in idx:
            v = values[idx["data_type"]] if idx["data_type"] < len(values) else ""
            if v:
                parts.append(f"Type={v}")
        if source_name:
            parts.insert(0, f"Source={source_name}")
        if parts:
            return " | ".join(parts)
    # Fallback: first 8 non-empty key=value
    context_parts = []
    for i, val in enumerate(values[:12]):
        if val:
            hdr = headers[i] if i < len(headers) else f"col_{i}"
            context_parts.append(f"{hdr}={val}")
    return " | ".join(context_parts[:8])


def _row_table_column(row: dict, doc_type: str | None) -> tuple[str | None, str | None]:
    """Return (table_name, column_name) when available from row."""
    headers = row.get("headers") or []
    values = row.get("values") or []
    idx = _build_header_index(headers, doc_type)
    table_name = None
    column_name = None
    if "table_name" in idx and idx["table_name"] < len(values):
        table_name = (values[idx["table_name"]] or "").strip() or None
    if "column_name" in idx and idx["column_name"] < len(values):
        column_name = (values[idx["column_name"]] or "").strip() or None
    return table_name, column_name


# ============================================================================
# Parsers
# ============================================================================

def _parse_xlsx(filepath: str) -> list[dict]:
    """Parse .xlsx file into list of {file, sheet, row_num, headers, values}."""
    if openpyxl is None:
        print("openpyxl is required for .xlsx files. Install: pip install openpyxl", file=sys.stderr)
        return []

    rows = []
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            str_values = [str(v) if v is not None else "" for v in row]
            if row_idx == 1:
                headers = str_values
                continue
            rows.append({
                "file": os.path.basename(filepath),
                "sheet": sheet_name,
                "row_num": row_idx,
                "headers": headers,
                "values": str_values,
            })
    wb.close()
    return rows


def _parse_xls(filepath: str) -> list[dict]:
    """Parse .xls file into list of dicts."""
    if xlrd is None:
        print("xlrd is required for .xls files. Install: pip install xlrd", file=sys.stderr)
        return []

    rows = []
    wb = xlrd.open_workbook(filepath)
    for sheet in wb.sheets():
        headers = []
        for row_idx in range(sheet.nrows):
            str_values = [str(sheet.cell_value(row_idx, c)) for c in range(sheet.ncols)]
            if row_idx == 0:
                headers = str_values
                continue
            rows.append({
                "file": os.path.basename(filepath),
                "sheet": sheet.name,
                "row_num": row_idx + 1,
                "headers": headers,
                "values": str_values,
            })
    return rows


def parse_file(filepath: str, use_cache: bool = True) -> list[dict]:
    """Parse an Excel file (with optional caching)."""
    if use_cache:
        cached = _load_cache(filepath)
        if cached is not None:
            return cached

    ext = Path(filepath).suffix.lower()
    if ext == ".xlsx":
        rows = _parse_xlsx(filepath)
    elif ext == ".xls":
        rows = _parse_xls(filepath)
    else:
        print(f"Unsupported file format: {ext}", file=sys.stderr)
        return []

    if use_cache:
        _save_cache(filepath, rows)

    return rows


# ============================================================================
# Search
# ============================================================================

def _file_matches_db(filepath: str, db: str | None) -> bool:
    """
    Check whether a documentation file should be searched for a given DB/system.

    db:
      - None       -> accept all files
      - "DWH"      -> only DWH docs (dwh-meta-*)
      - other name -> only source docs whose [source] matches db lowercased (pol, los, cif, ...)
    """
    if not db:
        return True

    db_norm = db.strip().lower()
    doc_type, source_name = _doc_type_from_filename(Path(filepath).name)

    if not doc_type:
        return False

    if db_norm == "dwh":
        return doc_type.startswith("dwh_")

    # For non-DWH aliases, match on source_name for source-system docs
    if source_name and doc_type.startswith("source_"):
        return source_name == db_norm

    return False


def search_documents(keyword: str, folder: str = "documents/",
                     use_cache: bool = True, use_regex: bool = False,
                     limit: int = 200, db: str | None = None) -> list[dict]:
    """
    Search all Excel files in folder for keyword.

    Returns list of matches with context: file, sheet, row_num, matched_field, context.
    """
    # Resolve folder relative to skill root
    skill_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    abs_folder = os.path.join(skill_root, folder) if not os.path.isabs(folder) else folder

    if not os.path.isdir(abs_folder):
        print(f"Thư mục không tồn tại: {abs_folder}", file=sys.stderr)
        return []

    # Collect all Excel files
    excel_files = []
    for f in os.listdir(abs_folder):
        if f.lower().endswith((".xlsx", ".xls")) and not f.startswith("~$"):
            fp = os.path.join(abs_folder, f)
            if _file_matches_db(fp, db):
                excel_files.append(fp)

    if not excel_files:
        print(f"Không tìm thấy file Excel trong: {abs_folder}", file=sys.stderr)
        return []

    results = []
    for fp in sorted(excel_files):
        rows = parse_file(fp, use_cache=use_cache)

        # Search in sheet names (reported once per sheet)
        seen_sheets = set()
        for row in rows:
            sheet = row["sheet"]
            if sheet not in seen_sheets and _match(keyword, sheet, use_regex):
                seen_sheets.add(sheet)
                _dt, _sn = _doc_type_from_filename(row["file"])
                entry = {
                    "file": row["file"], "sheet": sheet,
                    "row_num": None, "matched_field": "sheet_name",
                    "context": f"Sheet name: {sheet}",
                }
                if _dt:
                    entry["doc_type"] = _dt
                if _sn:
                    entry["source_name"] = _sn
                results.append(entry)
                seen_sheets.add(sheet)

        # Search in headers (reported once per sheet)
        seen_headers = set()
        for row in rows:
            sheet = row["sheet"]
            _dt, _sn = _doc_type_from_filename(row["file"])
            for hi, h in enumerate(row["headers"]):
                key = (sheet, hi)
                if key not in seen_headers and _match(keyword, h, use_regex):
                    seen_headers.add(key)
                    entry = {
                        "file": row["file"], "sheet": sheet,
                        "row_num": 1, "matched_field": "header",
                        "context": f"Header [{hi}]: {h}",
                    }
                    if _dt:
                        entry["doc_type"] = _dt
                    if _sn:
                        entry["source_name"] = _sn
                    results.append(entry)

        # Search in cell values
        for row in rows:
            _dt, _sn = _doc_type_from_filename(row["file"])
            for vi, v in enumerate(row["values"]):
                if _match(keyword, v, use_regex):
                    header = row["headers"][vi] if vi < len(row["headers"]) else f"col_{vi}"
                    # Use normalized context when we have a known doc type
                    ctx = _row_to_normalized_context(row, _dt, _sn)
                    tbl, col = _row_table_column(row, _dt)

                    entry = {
                        "file": row["file"], "sheet": row["sheet"],
                        "row_num": row["row_num"],
                        "matched_field": f"cell:{header}",
                        "context": ctx,
                    }
                    if _dt:
                        entry["doc_type"] = _dt
                    if _sn:
                        entry["source_name"] = _sn
                    if tbl:
                        entry["table_name"] = tbl
                    if col:
                        entry["column_name"] = col
                    results.append(entry)

            if len(results) >= limit:
                break
        if len(results) >= limit:
            break

    return results[:limit]


def _match(pattern: str, text: str, use_regex: bool) -> bool:
    if not text:
        return False
    if use_regex:
        return bool(re.search(pattern, text, re.IGNORECASE))
    return pattern.lower() in text.lower()


# ============================================================================
# Formatters
# ============================================================================

def format_text(results: list[dict]) -> str:
    if not results:
        return "Không tìm thấy kết quả nào."

    lines = [f"Tìm thấy {len(results)} kết quả:\n"]
    lines.append(f"{'FILE':<28} {'TYPE':<14} {'SHEET':<20} {'ROW':<5} {'MATCH':<12} {'CONTEXT'}")
    lines.append("-" * 140)
    for r in results:
        doc_type = r.get("doc_type") or ""
        if r.get("source_name"):
            doc_type = doc_type + ":" + r["source_name"]
        lines.append(
            f"{r['file']:<28} {doc_type:<14} {r['sheet']:<20} {str(r['row_num'] or ''):<5} "
            f"{r['matched_field']:<12} {(r['context'] or '')[:85]}"
        )
    return "\n".join(lines)


def format_json(results: list[dict]) -> str:
    return json.dumps(results, indent=2, ensure_ascii=False)


def format_markdown(results: list[dict]) -> str:
    if not results:
        return "Không tìm thấy kết quả nào."
    lines = [f"Tìm thấy **{len(results)}** kết quả:\n"]
    lines.append("| File | Type | Sheet | Row | Match | Context |")
    lines.append("|------|------|-------|-----|-------|---------|")
    for r in results:
        doc_type = r.get("doc_type") or ""
        if r.get("source_name"):
            doc_type = f"{doc_type}:{r['source_name']}"
        ctx = (r["context"] or "")[:75].replace("|", "\\|")
        lines.append(
            f"| {r['file']} | {doc_type} | {r['sheet']} | {r['row_num'] or ''} "
            f"| {r['matched_field']} | {ctx} |"
        )
    return "\n".join(lines)


# ============================================================================
# CLI
# ============================================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Search Excel documentation files for table/column information."
    )
    parser.add_argument("--keyword", "-k", required=True, help="Search keyword or regex")
    parser.add_argument("--folder", default="documents/", help="Folder containing Excel files")
    parser.add_argument("--no-cache", action="store_true", help="Disable caching")
    parser.add_argument("--regex", action="store_true", help="Use regex matching")
    parser.add_argument("--limit", type=int, default=200, help="Max results (default: 200)")
    parser.add_argument(
        "--db",
        help="Limit docs to a specific system: DWH (warehouse), or source alias like POL, LOS, CIF, ...",
    )
    parser.add_argument("--format", choices=["text", "json", "markdown"], default="text")
    args = parser.parse_args()

    try:
        results = search_documents(
            keyword=args.keyword, folder=args.folder,
            use_cache=not args.no_cache, use_regex=args.regex,
            limit=args.limit, db=args.db,
        )
        if args.format == "json":
            print(format_json(results))
        elif args.format == "markdown":
            print(format_markdown(results))
        else:
            print(format_text(results))
    except Exception as e:
        print(f"Lỗi: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)
